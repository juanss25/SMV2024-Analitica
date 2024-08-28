from flask import Flask, request, jsonify
import openpyxl

app = Flask(__name__)

# Variable global para almacenar los datos
data = None
@app.route('/')
def index():
    return jsonify({"hola mundo git"})
    
# Ruta para subir y procesar el archivo Excel
@app.route('/upload', methods=['POST'])
def upload_file():
    global data  # Usar la variable global
    file = request.files.get('file')

    if not file:
        return jsonify({"error": "No file uploaded"}), 400

    try:
        # Cargar el archivo Excel usando openpyxl
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        data = []
        for row in range(2, sheet.max_row + 1):
            row_data = {
                'ID': sheet.cell(row=row, column=1).value,
                'Nombre': sheet.cell(row=row, column=2).value,
                'Email': sheet.cell(row=row, column=3).value
            }
            data.append(row_data)

        return jsonify({"message": "File uploaded successfully", "data": data}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 400

# Ruta para buscar en los datos cargados
@app.route('/search', methods=['POST'])
def search():
    global data  # Usar la variable global

    if data is None:
        return jsonify({"error": "No data loaded. Please upload a file first."}), 400

    search_type = request.json.get('search_type')
    search_value = request.json.get('search_value')

    if not search_type or not search_value:
        return jsonify({"error": "search_type and search_value are required"}), 400

    try:
        if search_type == "Nombre":
            result = [row for row in data if search_value.lower() in row['Nombre'].lower()]
        elif search_type == "ID":
            result = [row for row in data if row['ID'] == int(search_value)]
        else:
            return jsonify({"error": "Invalid search type. Use 'Nombre' or 'ID'."}), 400

        return jsonify(result), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 400

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
